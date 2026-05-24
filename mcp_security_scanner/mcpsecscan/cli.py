from __future__ import annotations
import asyncio, csv, json
from collections import Counter, defaultdict
from hashlib import sha256
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZIP_DEFLATED, ZipFile
import typer, yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from .client import scan_stdio, scan_http, MCPClientError
from .models import Finding, ScanReport, ScanTarget
from .rules import run_rules

app = typer.Typer(help="Scan MCP servers for security vulnerabilities and misconfigurations.")
console = Console()

SEV_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
SARIF_LEVEL = {"critical": "error", "high": "error", "medium": "warning", "low": "note", "info": "note"}
FINDING_COLUMNS = [
    "target",
    "transport",
    "severity",
    "confidence",
    "category",
    "id",
    "title",
    "description",
    "cwe",
    "tool",
    "evidence",
    "recommendation",
]

async def _scan(target: ScanTarget) -> ScanReport:
    observations = {}
    if target.transport == "stdio":
        server_info, tools = await scan_stdio(target)
    else:
        server_info, tools, observations = await scan_http(target)
    findings = run_rules(target, tools, server_info, observations)
    return ScanReport(target=target, server_info=server_info, tools=tools, findings=findings)

def write_outputs(
    reports: list[ScanReport],
    json_out: str | None,
    csv_out: str | None,
    sarif_out: str | None,
    excel_out: str | None,
    zip_out: str | None,
):
    output_paths = {
        "json": json_out,
        "csv": csv_out,
        "sarif": sarif_out,
        "excel": excel_out,
    }

    if zip_out:
        with TemporaryDirectory(prefix="mcpsecscan-") as tmpdir:
            tmp_paths = {
                key: (str(Path(tmpdir) / Path(value).name) if value else None)
                for key, value in output_paths.items()
            }
            generated_files = generate_selected_outputs(reports, tmp_paths)
            write_zip_bundle(zip_out, generated_files)
    else:
        generate_selected_outputs(reports, output_paths)
    print_console_summary(reports)

def generate_selected_outputs(reports: list[ScanReport], output_paths: dict[str, str | None]) -> list[str]:
    payload = [r.model_dump() for r in reports]
    generated_files: list[str] = []

    json_out = output_paths.get("json")
    if json_out:
        Path(json_out).write_text(json.dumps(payload, indent=2), encoding="utf-8")
        generated_files.append(json_out)

    csv_out = output_paths.get("csv")
    if csv_out:
        write_csv(reports, csv_out)
        generated_files.append(csv_out)

    sarif_out = output_paths.get("sarif")
    if sarif_out:
        Path(sarif_out).write_text(json.dumps(to_sarif(reports), indent=2), encoding="utf-8")
        generated_files.append(sarif_out)

    excel_out = output_paths.get("excel")
    if excel_out:
        write_excel(reports, excel_out)
        generated_files.append(excel_out)

    return generated_files

def write_zip_bundle(zip_out: str, generated_files: list[str]):
    if not generated_files:
        return
    manifest_entries: list[dict[str, str | int]] = []
    with ZipFile(zip_out, mode="w", compression=ZIP_DEFLATED) as zip_file:
        for file_path in generated_files:
            path = Path(file_path)
            if path.exists():
                zip_file.write(path, arcname=path.name)
                file_bytes = path.read_bytes()
                manifest_entries.append({
                    "file": path.name,
                    "size": len(file_bytes),
                    "sha256": sha256(file_bytes).hexdigest(),
                })
        zip_file.writestr(
            "manifest.json",
            json.dumps({"generator": "mcpsecscan", "files": manifest_entries}, indent=2),
        )

def build_finding_rows(reports: list[ScanReport]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for r in reports:
        for f in r.findings:
            rows.append({
                "target": r.target.name,
                "transport": r.target.transport,
                "severity": f.severity,
                "confidence": f.confidence,
                "category": f.category,
                "id": f.id,
                "title": f.title,
                "description": f.description,
                "cwe": f.cwe or "",
                "tool": str(f.evidence.get("tool", "")),
                "evidence": json.dumps(f.evidence, ensure_ascii=False),
                "recommendation": f.recommendation,
            })
    return rows

def write_csv(reports: list[ScanReport], csv_out: str):
    rows = build_finding_rows(reports)
    with open(csv_out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=FINDING_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

def write_excel(reports: list[ScanReport], excel_out: str):
    wb = Workbook()
    summary = wb.active
    summary.title = "Risk Summary"

    total_findings = sum(len(r.findings) for r in reports)
    severity_counts = Counter(f.severity for r in reports for f in r.findings)

    summary["A1"] = "MCP Security Scan Summary"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A3"] = "Total Targets"
    summary["B3"] = len(reports)
    summary["A4"] = "Total Findings"
    summary["B4"] = total_findings

    summary["A6"] = "Severity"
    summary["B6"] = "Count"
    summary["A6"].font = Font(bold=True)
    summary["B6"].font = Font(bold=True)
    for idx, severity in enumerate(["critical", "high", "medium", "low", "info"], start=7):
        summary[f"A{idx}"] = severity.upper()
        summary[f"B{idx}"] = severity_counts[severity]

    summary["A13"] = "Target"
    summary["B13"] = "Transport"
    summary["C13"] = "Findings"
    summary["D13"] = "Critical"
    summary["E13"] = "High"
    summary["F13"] = "Medium"
    summary["G13"] = "Low"
    for cell in ("A13", "B13", "C13", "D13", "E13", "F13", "G13"):
        summary[cell].font = Font(bold=True)

    row_idx = 14
    for report in reports:
        counts = Counter(f.severity for f in report.findings)
        summary[f"A{row_idx}"] = report.target.name
        summary[f"B{row_idx}"] = report.target.transport
        summary[f"C{row_idx}"] = len(report.findings)
        summary[f"D{row_idx}"] = counts["critical"]
        summary[f"E{row_idx}"] = counts["high"]
        summary[f"F{row_idx}"] = counts["medium"]
        summary[f"G{row_idx}"] = counts["low"]
        row_idx += 1

    details = wb.create_sheet("Findings")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(FINDING_COLUMNS, start=1):
        cell = details.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    rows = build_finding_rows(reports)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(FINDING_COLUMNS, start=1):
            details.cell(row=row_idx, column=col_idx, value=row[col_name])

    for ws in (summary, details):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

    wb.save(excel_out)

def to_sarif(reports: list[ScanReport]) -> dict:
    all_findings = [f for r in reports for f in r.findings]
    rules: dict[str, Finding] = {}
    for f in all_findings:
        rules.setdefault(f.id, f)

    return {
        "version": "2.1.0",
        "$schema": "https://json.schemastore.org/sarif-2.1.0.json",
        "runs": [
            {
                "tool": {
                    "driver": {
                        "name": "mcpsecscan",
                        "semanticVersion": "0.1.0",
                        "informationUri": "https://modelcontextprotocol.io/",
                        "rules": [
                            {
                                "id": f.id,
                                "name": f.title,
                                "shortDescription": {"text": f.title},
                                "fullDescription": {"text": f.description},
                                "help": {"text": f.recommendation},
                                "properties": {
                                    "category": f.category,
                                    "severity": f.severity,
                                    "confidence": f.confidence,
                                    "cwe": f.cwe,
                                },
                            }
                            for f in sorted(rules.values(), key=lambda x: x.id)
                        ],
                    }
                },
                "automationDetails": {
                    "id": "mcpsecscan/default",
                },
                "results": [
                    sarif_result(r, f)
                    for r in reports
                    for f in sorted(r.findings, key=lambda x: (SEV_ORDER.get(x.severity, 99), x.category, x.id))
                ],
            }
        ],
    }

def sarif_result(report: ScanReport, finding: Finding) -> dict:
    tool_name = finding.evidence.get("tool")
    location_uri = f"mcp-targets/{report.target.name}.json"
    if tool_name:
        location_uri = f"mcp-targets/{report.target.name}/tools/{tool_name}.json"
    stable_key = f"{report.target.name}|{report.target.transport}|{finding.id}|{tool_name or '-'}|{finding.title}"
    fingerprint = sha256(stable_key.encode("utf-8")).hexdigest()
    return {
        "ruleId": finding.id,
        "level": SARIF_LEVEL.get(finding.severity, "warning"),
        "message": {"text": f"{finding.title} | Severity: {finding.severity} | Confidence: {finding.confidence}"},
        "partialFingerprints": {
            "primaryLocationLineHash": fingerprint,
        },
        "locations": [
            {
                "physicalLocation": {
                    "artifactLocation": {"uri": location_uri}
                }
            }
        ],
        "properties": {
            "target": report.target.name,
            "transport": report.target.transport,
            "category": finding.category,
            "severity": finding.severity,
            "confidence": finding.confidence,
            "cwe": finding.cwe,
            "description": finding.description,
            "evidence": finding.evidence,
            "recommendation": finding.recommendation,
        },
    }

def print_console_summary(reports: list[ScanReport]):
    table = Table(title="MCP Security Scan")
    table.add_column("Target")
    table.add_column("Transport")
    table.add_column("Tools", justify="right")
    table.add_column("Findings", justify="right")
    table.add_column("Critical", justify="right")
    table.add_column("High", justify="right")
    table.add_column("Medium", justify="right")
    table.add_column("Low", justify="right")
    for r in reports:
        counts = Counter(f.severity for f in r.findings)
        table.add_row(
            r.target.name, r.target.transport, str(len(r.tools)), str(len(r.findings)),
            str(counts["critical"]), str(counts["high"]), str(counts["medium"]), str(counts["low"]),
        )
    console.print(table)

    for r in reports:
        grouped: dict[str, list[Finding]] = defaultdict(list)
        for f in r.findings:
            grouped[f.category].append(f)

        for category in sorted(grouped.keys()):
            rows = grouped[category]
            cat_counts = Counter(f.severity for f in rows)
            console.print(Panel.fit(
                f"{r.target.name} / {category}  "
                f"critical={cat_counts['critical']} high={cat_counts['high']} medium={cat_counts['medium']} low={cat_counts['low']}",
                title="Finding Group",
            ))
            group_table = Table(show_header=True, header_style="bold")
            group_table.add_column("Severity")
            group_table.add_column("Confidence")
            group_table.add_column("Rule")
            group_table.add_column("Tool")
            group_table.add_column("Description")
            for f in sorted(rows, key=lambda x: (SEV_ORDER.get(x.severity, 99), x.id)):
                desc = f.description
                if len(desc) > 130:
                    desc = desc[:127] + "..."
                group_table.add_row(
                    f.severity.upper(),
                    f.confidence.upper(),
                    f.id,
                    str(f.evidence.get("tool", "-")),
                    desc,
                )
            console.print(group_table)

@app.command()
def stdio(
    name: str,
    command: str,
    out: str | None = typer.Option("report.json", "--out", help="Write full JSON report (legacy alias)."),
    json_out: str | None = typer.Option(None, "--json", help="Write full JSON report."),
    csv_out: str | None = typer.Option(None, "--csv", help="Write CSV findings report."),
    sarif_out: str | None = typer.Option(None, "--sarif", help="Write SARIF 2.1.0 report."),
    excel_out: str | None = typer.Option(None, "--excel", help="Write formatted Excel report."),
    zip_out: str | None = typer.Option(None, "--zip", help="Write zip bundle of generated output files."),
):
    """Scan a local MCP server launched over stdio."""
    report = asyncio.run(_scan(ScanTarget(name=name, transport="stdio", command=command)))
    write_outputs([report], json_out or out, csv_out, sarif_out, excel_out, zip_out)

@app.command()
def http(
    name: str,
    url: str,
    bearer: str | None = typer.Option(None, "--bearer", help="Bearer token for authenticated MCP endpoints."),
    out: str | None = typer.Option("report.json", "--out", help="Write full JSON report (legacy alias)."),
    json_out: str | None = typer.Option(None, "--json", help="Write full JSON report."),
    csv_out: str | None = typer.Option(None, "--csv", help="Write CSV findings report."),
    sarif_out: str | None = typer.Option(None, "--sarif", help="Write SARIF 2.1.0 report."),
    excel_out: str | None = typer.Option(None, "--excel", help="Write formatted Excel report."),
    zip_out: str | None = typer.Option(None, "--zip", help="Write zip bundle of generated output files."),
):
    """Scan a remote MCP HTTP endpoint."""
    report = asyncio.run(_scan(ScanTarget(name=name, transport="http", url=url, bearer=bearer)))
    write_outputs([report], json_out or out, csv_out, sarif_out, excel_out, zip_out)

@app.command(name="config")
def config_scan(
    file: str,
    out: str | None = typer.Option("report.json", "--out", help="Write full JSON report (legacy alias)."),
    json_out: str | None = typer.Option(None, "--json", help="Write full JSON report."),
    csv_out: str | None = typer.Option(None, "--csv", help="Write CSV findings report."),
    sarif_out: str | None = typer.Option(None, "--sarif", help="Write SARIF 2.1.0 report."),
    excel_out: str | None = typer.Option(None, "--excel", help="Write formatted Excel report."),
    zip_out: str | None = typer.Option(None, "--zip", help="Write zip bundle of generated output files."),
):
    """Scan multiple MCP targets from a YAML config file."""
    data = yaml.safe_load(Path(file).read_text())
    reports = []
    for item in data.get("servers", []):
        try:
            reports.append(asyncio.run(_scan(ScanTarget(**item))))
        except MCPClientError as e:
            console.print(f"[red]Failed {item.get('name')}: {e}[/red]")
    write_outputs(reports, json_out or out, csv_out, sarif_out, excel_out, zip_out)
